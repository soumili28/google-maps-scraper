"""
Export module for saving scraped Google Maps listings to CSV and professionally styled XLSX.
Includes post-export file integrity and record count validation.
"""

import os
import csv
import re
from datetime import datetime, timezone
from typing import List, Tuple, Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.models import BusinessListing
from src.exceptions import ExportValidationError


HEADERS = [
    ("Name", "name", 30),
    ("Rating", "rating", 10),
    ("Reviews Count", "reviews_count", 15),
    ("Category", "category", 25),
    ("Address", "address", 40),
    ("Phone", "phone", 18),
    ("Website", "website", 35),
    ("Google Maps URL", "place_url", 45),
    ("Latitude", "latitude", 14),
    ("Longitude", "longitude", 14),
    ("Status / Hours", "status_or_hours", 25),
    ("Scraped At (UTC)", "scraped_at", 22),
]


def sanitize_filename(name: str) -> str:
    """Sanitizes a string to be safely used in filenames."""
    cleaned = re.sub(r"[^\w\-]+", "_", name.strip())
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    return cleaned[:50]


def generate_export_paths(query: str, output_dir: str = "output") -> Tuple[str, str]:
    """Generates timestamped CSV and XLSX file paths."""
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    slug = sanitize_filename(query)
    base_name = f"google_maps_{slug}_{timestamp}"
    
    csv_path = os.path.abspath(os.path.join(output_dir, f"{base_name}.csv"))
    xlsx_path = os.path.abspath(os.path.join(output_dir, f"{base_name}.xlsx"))
    return csv_path, xlsx_path


def export_to_csv(listings: List[BusinessListing], file_path: str) -> str:
    """
    Exports listings to standard UTF-8 encoded CSV.
    """
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    header_titles = [h[0] for h in HEADERS]
    field_keys = [h[1] for h in HEADERS]

    with open(file_path, mode="w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(header_titles)
        for item in listings:
            item_dict = item.to_dict()
            writer.writerow([item_dict.get(k, "") for k in field_keys])

    return file_path


def export_to_xlsx(listings: List[BusinessListing], file_path: str) -> str:
    """
    Exports listings to a styled Excel (.xlsx) workbook with:
    - Formatted header row (navy fill, bold white text, center aligned)
    - Autofilter enabled
    - Frozen header pane (A2)
    - Auto-adjusted column widths with sensible min/max constraints
    - Text wrapping on address, category, and status columns
    - Clickable hyperlinks for Website and Google Maps Place URL
    - Thin borders and clean typography
    """
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Google Maps Leads"

    # Style definitions
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="Calibri", size=10, color="000000")
    link_font = Font(name="Calibri", size=10, color="0563C1", underline="single")
    
    thin_border = Border(
        left=Side(style="thin", color="D3D3D3"),
        right=Side(style="thin", color="D3D3D3"),
        top=Side(style="thin", color="D3D3D3"),
        bottom=Side(style="thin", color="D3D3D3"),
    )
    header_border = Border(
        left=Side(style="thin", color="1B365D"),
        right=Side(style="thin", color="1B365D"),
        top=Side(style="thin", color="1B365D"),
        bottom=Side(style="medium", color="1B365D"),
    )

    wrap_alignment = Alignment(vertical="center", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    default_alignment = Alignment(vertical="center")

    # Write Headers
    header_titles = [h[0] for h in HEADERS]
    field_keys = [h[1] for h in HEADERS]
    default_widths = {h[0]: h[2] for h in HEADERS}

    ws.append(header_titles)
    ws.row_dimensions[1].height = 28

    for col_idx in range(1, len(header_titles) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border

    # Write Data Rows
    for row_idx, item in enumerate(listings, start=2):
        ws.row_dimensions[row_idx].height = 22
        item_dict = item.to_dict()

        for col_idx, key in enumerate(field_keys, start=1):
            val = item_dict.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = default_alignment

            # Center align numbers, ratings, coordinates, and dates
            if key in ("rating", "reviews_count", "latitude", "longitude", "scraped_at"):
                cell.alignment = center_alignment
                cell.value = val

            # Text wrapping for longer content
            elif key in ("address", "category", "status_or_hours"):
                cell.alignment = wrap_alignment
                cell.value = val

            # Clickable hyperlinks for URLs
            elif key in ("website", "place_url") and val:
                url_str = str(val).strip()
                if url_str.startswith("http://") or url_str.startswith("https://"):
                    cell.value = url_str
                    cell.hyperlink = url_str
                    cell.font = link_font
                else:
                    cell.value = val
            else:
                cell.value = val

    # Freeze Header Pane
    ws.freeze_panes = "A2"

    # Enable Autofilter
    if len(listings) > 0:
        ws.auto_filter.ref = ws.dimensions

    # Adjust Column Widths
    for col_idx, col_name in enumerate(header_titles, start=1):
        col_letter = get_column_letter(col_idx)
        # Find maximum length of content in column
        max_len = len(str(col_name))
        for row_idx in range(2, len(listings) + 2):
            cell_val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            if len(cell_val) > max_len:
                max_len = len(cell_val)
        
        suggested_width = max(default_widths.get(col_name, 15), min(max_len + 3, 50))
        ws.column_dimensions[col_letter].width = suggested_width

    wb.save(file_path)
    return file_path


def validate_export(file_path: str, expected_count: int) -> Dict[str, Any]:
    """
    Validates exported CSV or XLSX file:
    - Verifies file exists and has non-zero size
    - Verifies file is readable
    - Verifies row count matches expected record count
    """
    if not os.path.exists(file_path):
        raise ExportValidationError(f"Export validation failed: File not found at '{file_path}'")
    
    file_size = os.path.getsize(file_path)
    if file_size == 0:
        raise ExportValidationError(f"Export validation failed: File '{file_path}' is empty (0 bytes)")

    ext = os.path.splitext(file_path)[1].lower()
    actual_records = 0

    if ext == ".csv":
        with open(file_path, mode="r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = list(reader)
            if not rows:
                raise ExportValidationError(f"CSV file '{file_path}' has no header or data rows.")
            # Subtract 1 for the header row
            actual_records = max(0, len(rows) - 1)

    elif ext == ".xlsx":
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                raise ExportValidationError(f"XLSX file '{file_path}' has no content.")
            actual_records = max(0, len(rows) - 1)
            wb.close()
        except Exception as e:
            raise ExportValidationError(f"Failed to read XLSX file '{file_path}': {e}")
    else:
        raise ExportValidationError(f"Unsupported file extension for validation: {ext}")

    if actual_records != expected_count:
        raise ExportValidationError(
            f"Record count mismatch in '{file_path}': expected {expected_count}, found {actual_records}."
        )

    return {
        "file_path": file_path,
        "file_size_bytes": file_size,
        "format": ext.replace(".", "").upper(),
        "record_count": actual_records,
        "status": "VALID",
    }
