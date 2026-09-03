"""Tests for CSV/XLSX export formatting and post-export validation."""

import os
import tempfile
import pytest
import openpyxl

from src.models import BusinessListing
from src.exporter import (
    export_to_csv,
    export_to_xlsx,
    validate_export,
    generate_export_paths,
)
from src.exceptions import ExportValidationError


@pytest.fixture
def sample_listings():
    return [
        BusinessListing(
            name="Austin Smiles Dental",
            rating=4.9,
            reviews_count=180,
            category="Dentist",
            address="100 Main St, Austin, TX 78701",
            phone="+1 512-555-0101",
            website="https://austinsmiles.example.com",
            place_url="https://maps.google.com/?cid=111",
            latitude=30.2672,
            longitude=-97.7431,
            status_or_hours="Open ⋅ Closes 6 PM",
        ),
        BusinessListing(
            name="Capital City Dentistry",
            rating=4.7,
            reviews_count=95,
            category="Cosmetic Dentist",
            address="200 Congress Ave, Austin, TX 78701",
            phone="+1 512-555-0202",
            website="https://capitalcitydentistry.example.com",
            place_url="https://maps.google.com/?cid=222",
            latitude=30.2680,
            longitude=-97.7420,
            status_or_hours="Closed ⋅ Opens 8 AM Mon",
        ),
    ]


def test_csv_export_and_validation(sample_listings):
    with tempfile.TemporaryDirectory() as tmpdir:
        csv_path = os.path.join(tmpdir, "test_leads.csv")
        exported = export_to_csv(sample_listings, csv_path)
        assert os.path.exists(exported)
        
        # Validate export
        val_result = validate_export(exported, expected_count=2)
        assert val_result["status"] == "VALID"
        assert val_result["record_count"] == 2
        assert val_result["format"] == "CSV"


def test_xlsx_export_formatting_and_validation(sample_listings):
    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = os.path.join(tmpdir, "test_leads.xlsx")
        exported = export_to_xlsx(sample_listings, xlsx_path)
        assert os.path.exists(exported)

        # Validate file
        val_result = validate_export(exported, expected_count=2)
        assert val_result["status"] == "VALID"
        assert val_result["record_count"] == 2
        assert val_result["format"] == "XLSX"

        # Verify Excel properties & formatting
        wb = openpyxl.load_workbook(exported)
        ws = wb.active

        # 1. Freeze Panes
        assert ws.freeze_panes == "A2"

        # 2. Autofilter
        assert ws.auto_filter.ref is not None

        # 3. Header styling
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.font.bold is True
        assert header_cell.fill.start_color.rgb == "001F497D"

        # 4. Hyperlinks for website & place_url
        website_cell = ws.cell(row=2, column=7)  # Website column
        assert website_cell.hyperlink is not None
        assert website_cell.hyperlink.target == "https://austinsmiles.example.com"
        assert website_cell.font.underline == "single"

        # 5. Column dimensions
        assert ws.column_dimensions["A"].width > 10

        wb.close()


def test_validation_errors(sample_listings):
    with tempfile.TemporaryDirectory() as tmpdir:
        csv_path = os.path.join(tmpdir, "test_err.csv")
        export_to_csv(sample_listings, csv_path)

        # Mismatched count
        with pytest.raises(ExportValidationError) as exc_info:
            validate_export(csv_path, expected_count=10)
        assert "mismatch" in str(exc_info.value).lower()

        # Non-existent file
        with pytest.raises(ExportValidationError) as exc_info:
            validate_export(os.path.join(tmpdir, "non_existent.csv"), expected_count=2)
        assert "not found" in str(exc_info.value).lower()


def test_generate_export_paths():
    with tempfile.TemporaryDirectory() as tmpdir:
        csv_p, xlsx_p = generate_export_paths("Dentists in Austin, TX", output_dir=tmpdir)
        assert csv_p.endswith(".csv")
        assert xlsx_p.endswith(".xlsx")
        assert "Dentists_in_Austin_TX" in csv_p
